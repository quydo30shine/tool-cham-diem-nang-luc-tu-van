#!/usr/bin/env python3
"""
Cập nhật danh sách nhân sự trong tool chấm điểm, KHÔNG xoá nhân sự cũ.

Cách dùng (mỗi tháng, sau khi tải file thống kê mới về cùng thư mục này):

    python3 cap_nhat_nhan_su.py                 # xem trước, chưa ghi gì
    python3 cap_nhat_nhan_su.py --ghi           # ghi vào file HTML
    python3 cap_nhat_nhan_su.py --ghi --day     # ghi + đẩy lên link GitHub cho salesup

Quy tắc gộp:
  - Nhân sự có trong file mới        -> thêm mới, hoặc cập nhật salon/vị trí/level.
  - Nhân sự chỉ còn trong file cũ    -> GIỮ LẠI, gắn cờ "off" (tool hiện nhãn "Danh sách cũ",
                                        xếp cuối kết quả tìm kiếm, vẫn chấm được).
  - Một người làm nhiều salon        -> gộp 1 dòng, salon ghi "202 DC / 264 LLQ".

Điểm đã chấm nằm trên máy từng salesup (localStorage), script này không đụng tới.
"""

import argparse, json, re, subprocess, sys, unicodedata
from pathlib import Path

# File xlsx xuất ra dùng Unicode tổ hợp (NFD): "là" trong file khác byte với "là" gõ tay,
# nhìn giống hệt nhau nhưng so sánh chuỗi sẽ trượt. Chuẩn hoá NFC cho mọi text đọc vào.
nfc = lambda s: unicodedata.normalize("NFC", str(s))

HERE = Path(__file__).resolve().parent
HTML = HERE / "tool-cham-diem-nang-luc-tu-van.html"

# Vị trí nhận từ dòng "Bộ lọc được áp dụng" trong ô A1 của file xlsx.
VITRI = [
    ("staff_position là Skinner", "Skinner"),
    ("staff_position là Stylist", "Stylist"),
    ("staff_position là Sup DV", "Supporter"),
]


def doc_xlsx(path):
    """Trả về [{i,n,s,v,l}] từ 1 file thống kê. Bỏ qua file không nhận diện được vị trí."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Thiếu thư viện openpyxl. Chạy: pip install openpyxl")

    ws = openpyxl.load_workbook(path, data_only=True).active
    a1 = nfc(ws["A1"].value or "")
    vitri = next((v for key, v in VITRI if key in a1), None)
    if not vitri:
        return None, None

    # Dòng tiêu đề là dòng có cả 'Salon' và 'Tên NV'
    head_row = None
    for r in range(1, 12):
        vals = [nfc(c.value or "").strip() for c in ws[r]]
        if "Salon" in vals and "Tên NV" in vals:
            head_row = r
            cot = {v: k for k, v in enumerate(vals)}
            break
    if head_row is None:
        return None, None

    ra = []
    for row in ws.iter_rows(min_row=head_row + 1, values_only=True):
        salon = row[cot["Salon"]]
        ten = row[cot["Tên NV"]]
        level = row[cot.get("Level", -1)] if "Level" in cot else None
        if not salon or not ten:
            continue
        m = re.match(r"^(.*)-(\d+)$", nfc(ten).strip())
        if not m:
            print(f"  [bỏ qua] không tách được ID từ tên: {ten!r}")
            continue
        ra.append({
            "i": m.group(2),
            "n": m.group(1).strip(),
            "s": nfc(salon).strip(),
            "v": vitri,
            "l": nfc(level).replace("Level ", "L").strip() if level else "",
        })
    return vitri, ra


def gop_trung(ds):
    """Cùng 1 ID xuất hiện ở nhiều salon -> gộp thành 1 dòng, salon nối bằng ' / '."""
    ra = {}
    for o in ds:
        cu = ra.get(o["i"])
        if cu:
            if o["s"] not in cu["s"].split(" / "):
                cu["s"] += " / " + o["s"]
        else:
            ra[o["i"]] = dict(o)
    return ra


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ghi", action="store_true", help="ghi thay đổi vào file HTML")
    ap.add_argument("--day", action="store_true", help="commit + push lên GitHub Pages")
    args = ap.parse_args()

    html = HTML.read_text(encoding="utf-8")
    m = re.search(r"/\*STAFF_START\*/(.*?)/\*STAFF_END\*/", html, re.S)
    if not m:
        sys.exit(f"Không tìm thấy vùng dữ liệu nhân sự trong {HTML.name}")
    cu = {o["i"]: {k: (nfc(v) if isinstance(v, str) else v) for k, v in o.items()}
          for o in json.loads(m.group(1))}

    # --- đọc các file xlsx trong thư mục ---
    moi_raw = []
    print("Đọc file thống kê:")
    for f in sorted(HERE.glob("*.xlsx")):
        if f.name.startswith("~$"):
            continue
        vitri, ds = doc_xlsx(f)
        if ds is None:
            print(f"  [bỏ qua] {f.name} — không nhận ra vị trí trong ô A1")
            continue
        print(f"  {f.name} -> {vitri}: {len(ds)} dòng")
        moi_raw += ds

    if not moi_raw:
        sys.exit("Không đọc được nhân sự nào. Đặt file .xlsx thống kê vào cùng thư mục này.")

    moi = gop_trung(moi_raw)

    # --- gộp cũ + mới ---
    them, doi, giu = [], [], []
    ket = {}
    for i, o in moi.items():
        o.pop("off", None)
        if i in cu:
            c = cu[i]
            if any(c.get(k, "") != o.get(k, "") for k in ("n", "s", "v", "l")) or c.get("off"):
                doi.append((c, o))
        else:
            them.append(o)
        ket[i] = o
    for i, c in cu.items():
        if i not in ket:
            c["off"] = 1
            ket[i] = c
            giu.append(c)

    ds = sorted(ket.values(), key=lambda x: (x.get("off", 0), x["s"], x["n"]))

    # --- báo cáo ---
    print(f"\nCũ: {len(cu)} nhân sự  ->  Mới: {len(ds)} nhân sự")
    print(f"  + Thêm mới      : {len(them)}")
    print(f"  ~ Cập nhật      : {len(doi)}  (đổi salon / vị trí / level)")
    print(f"  = Giữ lại (cũ)  : {len(giu)}  (không còn trong file mới, vẫn chấm được)")
    for o in them[:8]:
        print(f"      + {o['i']} {o['n']} — {o['v']} — {o['s']}")
    if len(them) > 8:
        print(f"      … và {len(them)-8} người nữa")
    for c, o in doi[:8]:
        thay = ", ".join(f"{k}: {c.get(k,'')!r}→{o.get(k,'')!r}"
                         for k in ("n", "s", "v", "l") if c.get(k, "") != o.get(k, ""))
        print(f"      ~ {o['i']} {o['n']} — {thay or 'quay lại danh sách mới'}")
    if len(doi) > 8:
        print(f"      … và {len(doi)-8} người nữa")
    for c in giu[:8]:
        print(f"      = {c['i']} {c['n']} — {c['v']} — {c['s']}")
    if len(giu) > 8:
        print(f"      … và {len(giu)-8} người nữa")

    if not args.ghi:
        print("\nMới chỉ xem trước. Thêm --ghi để ghi vào file, --ghi --day để đẩy luôn lên link.")
        return

    data = json.dumps(ds, ensure_ascii=False, separators=(",", ":"))
    HTML.write_text(
        html[:m.start(1)] + data + html[m.end(1):], encoding="utf-8")
    print(f"\nĐã ghi {len(ds)} nhân sự vào {HTML.name}")

    if args.day:
        repo = HERE / ".deploy"
        url = "https://github.com/quydo30shine/tool-cham-diem-nang-luc-tu-van.git"
        if not repo.exists():
            subprocess.run(["git", "clone", "-q", url, str(repo)], check=True)
        subprocess.run(["git", "-C", str(repo), "pull", "-q"], check=True)
        (repo / "index.html").write_text(HTML.read_text(encoding="utf-8"), encoding="utf-8")
        subprocess.run(["git", "-C", str(repo), "add", "index.html"], check=True)
        r = subprocess.run(["git", "-C", str(repo), "commit", "-qm",
                            f"Cập nhật danh sách nhân sự: +{len(them)} mới, ~{len(doi)} đổi, ={len(giu)} giữ"])
        if r.returncode != 0:
            print("Không có gì thay đổi để đẩy.")
            return
        subprocess.run(["git", "-C", str(repo), "push", "-q", "origin", "main"], check=True)
        print("Đã đẩy lên: https://quydo30shine.github.io/tool-cham-diem-nang-luc-tu-van/")
        print("Salesup tải lại trang là thấy danh sách mới. Điểm đã chấm không bị mất.")


if __name__ == "__main__":
    main()
